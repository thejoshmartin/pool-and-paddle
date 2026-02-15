#!/usr/bin/env python3
"""
Parse contractor finish selections Excel → JSON.

Usage: python scripts/parse-finishes.py

Reads:  MARTIN- FINISHINGS.xlsx (place in project root)
Writes: src/finishes-data.json

This script was used to generate the initial finishes-data.json.
Re-run if the contractor updates the Excel template.
Requires: pip install openpyxl
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "MARTIN- FINISHINGS.xlsx"
OUTPUT_PATH = ROOT / "src" / "finishes-data.json"

CATEGORIES = [
    {"id": "flooring", "label": "Flooring", "icon": "🪵"},
    {"id": "shower-bath-tile", "label": "Shower/Bath Tile", "icon": "🚿"},
    {"id": "kitchens", "label": "Kitchens", "icon": "🍳"},
    {"id": "countertops", "label": "Countertops", "icon": "🪨"},
    {"id": "paint", "label": "Paint", "icon": "🎨"},
    {"id": "decking", "label": "Decking", "icon": "🏖️"},
    {"id": "doors", "label": "Doors", "icon": "🚪"},
    {"id": "plumbing", "label": "Plumbing", "icon": "🚰"},
    {"id": "appliances", "label": "Appliances", "icon": "🧊"},
    {"id": "electrical", "label": "Electrical", "icon": "💡"},
    {"id": "drywall", "label": "Drywall", "icon": "🏗️"},
]

ROOMS = [
    {"id": "whole-house", "label": "Whole House"},
    {"id": "master-suite", "label": "Master Suite"},
    {"id": "bed1-bath", "label": "Bedroom #1 + Bath"},
    {"id": "bed2", "label": "Bedroom #2"},
    {"id": "bed3-bath", "label": "Bedroom #3 + Bath"},
    {"id": "bunk-bath", "label": "Bunk Room + Bath"},
    {"id": "pool-bath", "label": "Pool Bath"},
    {"id": "upper-half-bath", "label": "Upper Half Bath"},
    {"id": "kitchen", "label": "Kitchen"},
    {"id": "wet-bar", "label": "Wet Bar"},
    {"id": "summer-kitchen", "label": "Summer Kitchen"},
    {"id": "laundry", "label": "Laundry"},
    {"id": "garage", "label": "Garage"},
    {"id": "exterior", "label": "Exterior"},
]

def parse_excel():
    if not EXCEL_PATH.exists():
        print(f"Excel file not found at: {EXCEL_PATH}")
        print("Place the contractor's Excel file in the project root and re-run.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    items = []
    item_id = 0

    # Category keywords to match sheet names / sections
    cat_keywords = {
        "floor": "flooring",
        "shower": "shower-bath-tile",
        "tile": "shower-bath-tile",
        "bath tile": "shower-bath-tile",
        "kitchen": "kitchens",
        "cabinet": "kitchens",
        "counter": "countertops",
        "paint": "paint",
        "deck": "decking",
        "door": "doors",
        "plumb": "plumbing",
        "faucet": "plumbing",
        "applian": "appliances",
        "electr": "electrical",
        "light": "electrical",
        "drywall": "drywall",
    }

    room_keywords = {
        "master": "master-suite",
        "bed 1": "bed1-bath",
        "bedroom 1": "bed1-bath",
        "bed 2": "bed2",
        "bedroom 2": "bed2",
        "bed 3": "bed3-bath",
        "bedroom 3": "bed3-bath",
        "bunk": "bunk-bath",
        "pool bath": "pool-bath",
        "half bath": "upper-half-bath",
        "kitchen": "kitchen",
        "wet bar": "wet-bar",
        "bar": "wet-bar",
        "summer": "summer-kitchen",
        "outdoor": "summer-kitchen",
        "laundry": "laundry",
        "garage": "garage",
        "exterior": "exterior",
        "outside": "exterior",
        "whole": "whole-house",
        "throughout": "whole-house",
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_category = None
        current_contractor_note = None

        # Try to detect category from sheet name
        sheet_lower = sheet_name.lower()
        for kw, cat_id in cat_keywords.items():
            if kw in sheet_lower:
                current_category = cat_id
                break

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or all(c is None for c in row):
                continue

            cells = [str(c).strip() if c is not None else "" for c in row]
            line = " | ".join(c for c in cells if c)

            if not line:
                continue

            # Detect category headers
            for kw, cat_id in cat_keywords.items():
                if kw in line.lower() and len(line) < 60:
                    current_category = cat_id
                    break

            # Detect room from cell content
            detected_room = "whole-house"
            for kw, room_id in room_keywords.items():
                if kw in line.lower():
                    detected_room = room_id
                    break

            # If row looks like an item (has descriptive text, not a header)
            if current_category and len(cells[0]) > 3 and not cells[0].isupper():
                item_id += 1
                contractor_opts = [c for c in cells[1:4] if c and len(c) > 2]

                items.append({
                    "id": f"x{item_id}",
                    "category": current_category,
                    "room": detected_room,
                    "item": cells[0],
                    "contractorOptions": contractor_opts,
                    "selection": "",
                    "unitPrice": None,
                    "quantity": None,
                    "unit": "ea",
                    "url": "",
                    "notes": "",
                })

    return items


def main():
    if EXCEL_PATH.exists():
        items = parse_excel()
        print(f"Parsed {len(items)} items from Excel")
    else:
        print(f"No Excel file found at {EXCEL_PATH}")
        print("Using existing finishes-data.json (already committed)")
        return

    output = {
        "categories": CATEGORIES,
        "rooms": ROOMS,
        "items": items,
    }

    # Add contractor notes (these are manually curated)
    for cat in output["categories"]:
        cat["contractorNote"] = ""

    OUTPUT_PATH.write_text(json.dumps(output, indent=2, ensure_ascii=False))
    print(f"Written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
